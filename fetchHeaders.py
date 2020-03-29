from openpyxl import load_workbook
import csv


def format_data(sheet, spreadsheet):
    ws = spreadsheet[sheet]
    max_column = 0
    sheet_content = []
    row_content = []
    for cell in ws[1]:
        if cell.value:
            row_content.append(cell.value.strip())
            max_column = cell.column
    sheet_content.append(row_content)

    for row in ws.iter_rows(min_row=2, max_col=max_column):
        row_content = []
        for cell in row:
            cell_value = cell.value
            if cell.data_type == 's':
                cell_value = cell.value.strip()
            row_content.append(cell_value)
        sheet_content.append(row_content)
    return sheet_content


def save_titles(spreadsheet, sheet_name, csv_file):
    file = open(csv_file, 'w')
    filewriter = csv.writer(file, delimiter=',', quoting=csv.QUOTE_MINIMAL)
    sheet_data = format_data(sheet_name, spreadsheet)
    for sheet_row in sheet_data:
        filewriter.writerow(sheet_row)
    return


columns_sheet = load_workbook('columnNames.xlsx')
save_titles(columns_sheet, 'Sheet1', 'columnNames.csv')
