from openpyxl import load_workbook
import csv

columns = []
total_rows = 0
total_cells = 0
parent_columns = ['Food Name', 'Food Code', 'Description', 'Group', 'Previous', 'Main data references', 'Footnote']
excel_file = 'CoFID_2019.xlsx'

print('Processing file "' + excel_file + '"...')

workbook = load_workbook(excel_file)
sheets_quantity = len(workbook.sheetnames)
parent_names = ['table_list', 'notes', 'factors', 'proximates', 'inorganics', 'vitamins', 'vitamin_fractions', 'SFA_FA',
					 'SFA', 'MUFA_FA', 'MUFA', 'PUFA_FA', 'PUFA', 'phytosterols', 'organic_acids']

print('Processing ' + str(sheets_quantity) + ' worksheets...')

with open('columnNames.csv', 'r') as f:
	reader = csv.reader(f)
	next(reader)
	for row in reader:
		columns.append(row)


def format_data(sheet, parent):
	global total_rows
	global total_cells
	excel_sheet = workbook[sheet]
	max_column = 0
	sheet_content = []
	nutrients_values = []
	header_row_cells = []

	# Process the first row (header)
	for column_index, cell in enumerate(excel_sheet[1]):
		if cell.value:
			for field in columns:
				cell_value = cell.value.strip()
				if cell_value == field[0].strip():
					if cell_value in parent_columns or sheet in ['List of tables', '1.1 Notes', '1.2 Factors']:
						header_row_cells.append(field[2].strip())
					else:
						# header_row_cells.append(parent + '.' + field[2].strip())
						nutrients_values.append({'index': column_index, 'label': field[1].strip(), 'unit': field[3].strip()})
						header_row_cells.append(parent + '.' + field[2].strip() + '.label')
						header_row_cells.append(parent + '.' + field[2].strip() + '.unit')
						header_row_cells.append(parent + '.' + field[2].strip() + '.quantity')
			max_column = cell.column
	sheet_content.append(header_row_cells)

	# Process remaining rows
	for excel_row in excel_sheet.iter_rows(min_row=4, max_col=max_column):
		row_content = []
		for column_index, cell in enumerate(excel_row):
			cell_value = cell.value
			if cell.data_type == 's':
				cell_value = cell.value.strip()
			for nutrient in nutrients_values:
				if column_index == nutrient['index']:
					row_content.append(nutrient['label'])
					row_content.append(nutrient['unit'])
					total_cells += 2
					break

			row_content.append(cell_value)
			total_cells += 1
		total_rows += 1
		sheet_content.append(row_content)
	return sheet_content


def format_number(number):
	return '%.0f' % number


# Cycles through all worksheets within the workbook
for index, sheet_name in enumerate(workbook.sheetnames):
	newSheetName = sheet_name.replace('.', '_').replace(' ', '_').replace('(', '').replace(')', '')
	file = open('./csv_exports/cofid_' + newSheetName + '.csv', 'w')
	filewriter = csv.writer(file, delimiter=',', quoting=csv.QUOTE_MINIMAL)

	sheet_data = format_data(sheet_name, parent_names[index])
	for sheet_row in sheet_data:
		filewriter.writerow(sheet_row)

	# Print progress to console
	print('...' + str(format_number((index + 1) * 100 / sheets_quantity) + '%'), end='')

# Add thousands separator to totals
total_rows = str(f'{total_rows:,}')
total_cells = str(f'{total_cells:,}')

print()
print('Completed processing ' + total_rows + ' rows (' + total_cells + ' cells) of data!!')
